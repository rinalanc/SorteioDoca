import streamlit as st
import random
import collections
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io # Para manipular arquivos em memória
import pandas as pd # Para exibir DataFrames no Streamlit

# --- Configuration Data ---
# Core positions for Before/After Dinner shifts (fixed positions in B/C columns)
CORE_POSITIONS = [
    "Recirculação 1", "Recirculação 2", "Recirculação 3",
    "Pesado 1", "Pesado 2",
    "Auditoria 1", "Auditoria 2", "Auditoria 3",
    "Azul 1", "Azul 2",
    "Shuttle 1", "Shuttle 2", "Shuttle 3",
    "Carregamento 1", "Carregamento 2",
    "GAP 1", "GAP 2"
]

# Activated functions for drawing (sporadic, allocated in G column) - ATUALIZADO
ACTIVATED_FUNCTIONS = [
    "Hatae - tirar pacote", "Triagem 1", "Triagem 2", "Triagem 3", "Triagem 4",
    "Bipando Hatae 1", "Bipando Hatae 2", # Dividido em duas vagas
    "Curva - Tirar Pacote", # Nova posição
    "Triagem Hatae 1", "Triagem Hatae 2", "Triagem Hatae 3", "Triagem Hatae 4", # Novas posições
    "Curva 2 - Tirar Pacote" # Nova posição
]

# Mapeamento de posições conceituais para aplicar a regra "não pode estar em conceituais iguais"
# ATUALIZADO para incluir as novas funções extras
CONCEPTUAL_POSITION_GROUPS = {
    "Shuttle": ["Shuttle 1", "Shuttle 2", "Shuttle 3"],
    "Auditoria": ["Auditoria 1", "Auditoria 2", "Auditoria 3"],
    "Recirculação": ["Recirculação 1", "Recirculação 2", "Recirculação 3"],
    "Pesado": ["Pesado 1", "Pesado 2"],
    "Azul": ["Azul 1", "Azul 2"],
    "GAP": ["GAP 1", "GAP 2"],
    "Carregamento": ["Carregamento 1", "Carregamento 2"],
    # Novas entradas para as funções ativadas
    "Triagem": ["Triagem 1", "Triagem 2", "Triagem 3", "Triagem 4", "Triagem Hatae 1", "Triagem Hatae 2", "Triagem Hatae 3", "Triagem Hatae 4"],
    "Hatae": ["Hatae - tirar pacote", "Bipando Hatae 1", "Bipando Hatae 2"], # "Bipando Hatae" agora faz parte do grupo "Hatae"
    "Curva": ["Curva - Tirar Pacote", "Curva 2 - Tirar Pacote"] # Adicionado Curva 2
}

_INVERTED_CONCEPTUAL_GROUPS = {}
for group, positions in CONCEPTUAL_POSITION_GROUPS.items():
    for pos in positions:
        _INVERTED_CONCEPTUAL_GROUPS[pos] = group

ALLOWED_IN_AZUL = ["rinalanc", "leonarsd", "horaroge", "silvnpau", "sousthib"]

# As regras de EXCLUSIONS e INCREASED_PROBABILITY agora devem considerar tanto a posição exata
# quanto o grupo conceitual da posição, conforme definido em CONCEPTUAL_POSITION_GROUPS.
# A lógica de choose_associate_with_rules será ajustada para isso.

EXCLUSIONS = {
    "rinalanc": {
        "GeneralCeia": ["Shuttle 1", "Shuttle 2", "Shuttle 3", "Recirculação 1", "Recirculação 2", "Recirculação 3", "Pesado 1", "Pesado 2", "GAP 1", "GAP 2"],
        "GeneralDraw": ["Shuttle 1", "Shuttle 2", "Shuttle 3", "Recirculação 1", "Recirculação 2", "Recirculação 3",
                        "Pesado 1", "Pesado 2", "GAP 1", "GAP 2", # Posições de Ceia
                        "Hatae - tirar pacote", "Triagem 1", "Triagem 2", "Triagem 3", "Triagem 4", # Funções Ativadas
                        "Bipando Hatae 1", "Bipando Hatae 2", "Curva - Tirar Pacote",
                        "Triagem Hatae 1", "Triagem Hatae 2", "Triagem Hatae 3", "Triagem Hatae 4", "Curva 2 - Tirar Pacote"] # Novas Funções Ativadas
    },
    "leonarsd": {
        "GeneralCeia": ["Shuttle 1", "Shuttle 2", "Shuttle 3", "Recirculação 1", "Recirculação 2", "Recirculação 3", "Pesado 1", "Pesado 2", "GAP 1", "GAP 2"],
        "GeneralDraw": ["Shuttle 1", "Shuttle 2", "Shuttle 3", "Recirculação 1", "Recirculação 2", "Recirculação 3",
                        "Pesado 1", "Pesado 2", "GAP 1", "GAP 2",
                        "Hatae - tirar pacote", "Triagem 1", "Triagem 2", "Triagem 3", "Triagem 4",
                        "Bipando Hatae 1", "Bipando Hatae 2", "Curva - Tirar Pacote",
                        "Triagem Hatae 1", "Triagem Hatae 2", "Triagem Hatae 3", "Triagem Hatae 4", "Curva 2 - Tirar Pacote"]
    },
    "horaroge": {
        "GeneralCeia": ["Pesado 1", "Pesado 2"],
        "GeneralDraw": ["Pesado 1", "Pesado 2"]
    },
    "silvnpau": {
        # Sem exclusões específicas.
    },
    "sousthib": {
        # Sem exclusões específicas.
    },
    "ferrlucq": {
        "GeneralCeia": ["Azul 1", "Azul 2", "Carregamento 1", "Carregamento 2"],
        "GeneralDraw": ["Azul 1", "Azul 2", "Carregamento 1", "Carregamento 2"]
    },
    "ksilsilv": {
        "GeneralCeia": ["Azul 1", "Azul 2", "Carregamento 1", "Carregamento 2"],
        "GeneralDraw": ["Azul 1", "Azul 2", "Carregamento 1", "Carregamento 2"]
    },
    "wessouzf": {
        "GeneralCeia": ["Azul 1", "Azul 2", "Carregamento 1", "Carregamento 2"],
        "GeneralDraw": ["Azul 1", "Azul 2", "Carregamento 1", "Carregamento 2"]
    },
    "piluanaq": {
        # Sem exclusões específicas.
    },
    "pretojon": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "EVAWWELI": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "rabsouza": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "lucenama": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "pedrour": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "ferrlnat": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "doubsant": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "vinichda": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "hjosesil": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "tmarcoso": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "luizsanp": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "nasckluc": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "salucasi": {
        "GeneralCeia": [],
        "GeneralDraw": []
    },
    "mlucneri": {
        "GeneralCeia": [],
        "GeneralDraw": []
    }
}

INCREASED_PROBABILITY = {
    "rinalanc": {
        "GeneralCeia": { "Azul 1": 4, "Azul 2": 4, "Auditoria 1": 3, "Auditoria 2": 3, "Auditoria 3": 3, "Carregamento 1": 3, "Carregamento 2": 3 },
        "GeneralDraw": {
            "Azul 1": 4, "Azul 2": 4, "Auditoria 1": 3, "Auditoria 2": 3, "Auditoria 3": 3, "Carregamento 1": 4, "Carregamento 2": 4,
            "Hatae - tirar pacote": 1, "Triagem 1": 1, "Triagem 2": 1, "Triagem 3": 1, "Triagem 4": 1,
            "Bipando Hatae 1": 1, "Bipando Hatae 2": 1, "Curva - Tirar Pacote": 1,
            "Triagem Hatae 1": 1, "Triagem Hatae 2": 1, "Triagem Hatae 3": 1, "Triagem Hatae 4": 1, "Curva 2 - Tirar Pacote": 1
        }
    },
    "leonarsd": {
        "GeneralCeia": { "Azul 1": 4, "Azul 2": 4, "Auditoria 1": 3, "Auditoria 2": 3, "Auditoria 3": 3, "Carregamento 1": 3, "Carregamento 2": 3 },
        "GeneralDraw": {
            "Azul 1": 4, "Azul 2": 4, "Auditoria 1": 3, "Auditoria 2": 3, "Auditoria 3": 3, "Carregamento 1": 4, "Carregamento 2": 4,
            "Hatae - tirar pacote": 1, "Triagem 1": 1, "Triagem 2": 1, "Triagem 3": 1, "Triagem 4": 1,
            "Bipando Hatae 1": 1, "Bipando Hatae 2": 1, "Curva - Tirar Pacote": 1,
            "Triagem Hatae 1": 1, "Triagem Hatae 2": 1, "Triagem Hatae 3": 1, "Triagem Hatae 4": 1, "Curva 2 - Tirar Pacote": 1
        }
    },
    "horaroge": {
        "GeneralCeia": { "Azul 1": 1, "Azul 2": 1, "Carregamento 1": 2, "Carregamento 2": 2, "Auditoria 1": 2, "Auditoria 2": 2, "Auditoria 3": 2, "Recirculação 1": 1, "Recirculação 2": 1, "Recirculação 3": 1, "GAP 1": 1, "GAP 2": 1, "Shuttle 1": 1, "Shuttle 2": 1, "Shuttle 3": 1 },
        "GeneralDraw": {
            "Azul 1": 4, "Azul 2": 4, "Carregamento 1": 2, "Carregamento 2": 2,
            "Hatae - tirar pacote": 1, "Triagem 1": 1, "Triagem 2": 1, "Triagem 3": 1, "Triagem 4": 1,
            "Bipando Hatae 1": 1, "Bipando Hatae 2": 1, "Curva - Tirar Pacote": 1,
            "Triagem Hatae 1": 1, "Triagem Hatae 2": 1, "Triagem Hatae 3": 1, "Triagem Hatae 4": 1, "Curva 2 - Tirar Pacote": 1
        }
    },
    "silvnpau": {
        "GeneralCeia": { "Azul 1": 1, "Azul 2": 1, "Carregamento 1": 2, "Carregamento 2": 2, "Recirculação 1": 1, "Recirculação 2": 1, "Recirculação 3": 1, "Pesado 1": 1, "Pesado 2": 1, "Auditoria 1": 1, "Auditoria 2": 1, "Auditoria 3": 1, "Shuttle 1": 1, "Shuttle 2": 1, "Shuttle 3": 1, "GAP 1": 1, "GAP 2": 1 },
        "GeneralDraw": {
            "Azul 1": 4, "Azul 2": 4, "Carregamento 1": 2, "Carregamento 2": 2,
            "Hatae - tirar pacote": 1, "Triagem 1": 1, "Triagem 2": 1, "Triagem 3": 1, "Triagem 4": 1,
            "Bipando Hatae 1": 1, "Bipando Hatae 2": 1, "Curva - Tirar Pacote": 1,
            "Triagem Hatae 1": 1, "Triagem Hatae 2": 1, "Triagem Hatae 3": 1, "Triagem Hatae 4": 1, "Curva 2 - Tirar Pacote": 1
        }
    },
    "sousthib": {
        "GeneralCeia": { "Azul 1": 1, "Azul 2": 1, "Carregamento 1": 2, "Carregamento 2": 2, "Pesado 1": 2, "Pesado 2": 2, "Auditoria 1": 1, "Auditoria 2": 1, "Auditoria 3": 1, "Recirculação 1": 1, "Recirculação 2": 1, "Recirculação 3": 1, "GAP 1": 1, "GAP 2": 1, "Shuttle 1": 1, "Shuttle 2": 1, "Shuttle 3": 1 },
        "GeneralDraw": {
            "Azul 1": 4, "Azul 2": 4, "Carregamento 1": 2, "Carregamento 2": 2, "Pesado 1": 2, "Pesado 2": 2,
            "Hatae - tirar pacote": 1, "Triagem 1": 1, "Triagem 2": 1, "Triagem 3": 1, "Triagem 4": 1,
            "Bipando Hatae 1": 1, "Bipando Hatae 2": 1, "Curva - Tirar Pacote": 1,
            "Triagem Hatae 1": 1, "Triagem Hatae 2": 1, "Triagem Hatae 3": 1, "Triagem Hatae 4": 1, "Curva 2 - Tirar Pacote": 1
        }
    },
    "ferrlucq": {
        "GeneralCeia": {
            "GAP 1": 2, "GAP 2": 2, "Shuttle 1": 2, "Shuttle 2": 2, "Shuttle 3": 2,
            "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2, "Pesado 1": 2, "Pesado 2": 2
        },
        "GeneralDraw": {
            "Pesado 1": 2, "Pesado 2": 2, "GAP 1": 2, "GAP 2": 2, "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2,
            "Hatae - tirar pacote": 2, "Triagem 1": 2, "Triagem 2": 2,
            "Triagem 3": 2, "Triagem 4": 2, "Bipando Hatae 1": 2, "Bipando Hatae 2": 2, "Curva - Tirar Pacote": 2,
            "Triagem Hatae 1": 2, "Triagem Hatae 2": 2, "Triagem Hatae 3": 2, "Triagem Hatae 4": 2, "Curva 2 - Tirar Pacote": 2,
            "Shuttle 1": 2, "Shuttle 2": 2, "Shuttle 3": 2
        }
    },
    "ksilsilv": {
        "GeneralCeia": {
            "GAP 1": 2, "GAP 2": 2, "Shuttle 1": 2, "Shuttle 2": 2, "Shuttle 3": 2,
            "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2, "Pesado 1": 2, "Pesado 2": 2
        },
        "GeneralDraw": {
            "Pesado 1": 2, "Pesado 2": 2, "GAP 1": 2, "GAP 2": 2, "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2,
            "Hatae - tirar pacote": 2, "Triagem 1": 2, "Triagem 2": 2,
            "Triagem 3": 2, "Triagem 4": 2, "Bipando Hatae 1": 2, "Bipando Hatae 2": 2, "Curva - Tirar Pacote": 2,
            "Triagem Hatae 1": 2, "Triagem Hatae 2": 2, "Triagem Hatae 3": 2, "Triagem Hatae 4": 2, "Curva 2 - Tirar Pacote": 2,
            "Shuttle 1": 2, "Shuttle 2": 2, "Shuttle 3": 2
        }
    },
    "pretojon": {
        "GeneralDraw": { "Azul 1": 3, "Azul 2": 3, "Auditoria 1": 2, "Auditoria 2": 2, "Auditoria 3": 2, "Carregamento 1": 2, "Carregamento 2": 2 }
    },
    "wessouzf": {
        "GeneralCeia": { "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2, "GAP 1": 3, "GAP 2": 3, "Auditoria 1": 1, "Auditoria 2": 1, "Auditoria 3": 1, "Shuttle 1": 1, "Shuttle 2": 1, "Shuttle 3": 1, "Pesado 1": 1, "Pesado 2": 1 },
        "GeneralDraw": {
            "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2, "GAP 1": 3, "GAP 2": 3,
            "Hatae - tirar pacote": 1, "Triagem 1": 1, "Triagem 2": 1, "Triagem 3": 1, "Triagem 4": 1,
            "Bipando Hatae 1": 1, "Bipando Hatae 2": 1, "Curva - Tirar Pacote": 1,
            "Triagem Hatae 1": 1, "Triagem Hatae 2": 1, "Triagem Hatae 3": 1, "Triagem Hatae 4": 1, "Curva 2 - Tirar Pacote": 1,
            "Auditoria 1": 1, "Auditoria 2": 1, "Auditoria 3": 1, "Shuttle 1": 1, "Shuttle 2": 1, "Shuttle 3": 1, "Pesado 1": 1, "Pesado 2": 1
        }
    },
    "piluanaq": {
        "GeneralCeia": { "Bipando Hatae 1": 2, "Bipando Hatae 2": 2, "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2, "Auditoria 1": 2, "Auditoria 2": 2, "Auditoria 3": 2, "GAP 1": 2, "GAP 2": 2, "Pesado 1": 1, "Pesado 2": 1, "Shuttle 1": 1, "Shuttle 2": 1, "Shuttle 3": 1, "Carregamento 1": 1, "Carregamento 2": 1 },
        "GeneralDraw": {
            "Bipando Hatae 1": 2, "Bipando Hatae 2": 2, "Recirculação 1": 2, "Recirculação 2": 2, "Recirculação 3": 2, "Auditoria 1": 2, "Auditoria 2": 2, "Auditoria 3": 2, "GAP 1": 2, "GAP 2": 2, "Pesado 1": 1, "Pesado 2": 1, "Shuttle 1": 1, "Shuttle 2": 1, "Shuttle 3": 1, "Carregamento 1": 1, "Carregamento 2": 1,
            "Hatae - tirar pacote": 1, "Triagem 1": 1, "Triagem 2": 1, "Triagem 3": 1, "Triagem 4": 1, "Curva - Tirar Pacote": 1,
            "Triagem Hatae 1": 1, "Triagem Hatae 2": 1, "Triagem Hatae 3": 1, "Triagem Hatae 4": 1, "Curva 2 - Tirar Pacote": 1
        }
    }
}

CORE_ASSOCIATES_FOR_DINNER = ["rinalanc", "leonarsd"]

# --- Helper Function for Weighted Random Choice with Rules ---
def choose_associate_with_rules(available_associates_in_this_round, position_name, time_slot_context, exclusions, probabilities, additional_exclusions_for_assoc=None):
    
    eligible_associates = []
    weights = []

    for assoc in available_associates_in_this_round:
        current_weight = 1
        is_eligible = True # Assume eligible until an exclusion is found

        # Get conceptual group for the position, if any
        position_conceptual_group = _INVERTED_CONCEPTUAL_GROUPS.get(position_name, None)

        # 1. IMMEDIATE EXCLUSIONS (HIGHEST PRIORITY)

        # Check general exclusion for Azul if associate is not in ALLOWED_IN_AZUL
        if position_name.startswith("Azul") and assoc not in ALLOWED_IN_AZUL:
            is_eligible = False
        
        # Check explicit exclusions from the EXCLUSIONS dictionary
        if is_eligible and assoc in exclusions:
            assoc_exclusions = exclusions[assoc]
            
            # Check if current position_name is in the specific time_slot_context's exclusions
            if time_slot_context in assoc_exclusions and position_name in assoc_exclusions[time_slot_context]:
                is_eligible = False
            # Check if conceptual group is in the specific time_slot_context's exclusions
            elif is_eligible and position_conceptual_group and time_slot_context in assoc_exclusions:
                if any(p in assoc_exclusions[time_slot_context] for p in CONCEPTUAL_POSITION_GROUPS.get(position_conceptual_group, [])):
                    is_eligible = False
            
            # If not excluded by specific context, check if it's in the general context (e.g., GeneralCeia/GeneralDraw)
            if is_eligible: # Only check if not already excluded by a more specific rule
                # Apply GeneralCeia rules for AntesCeia and DepoisCeia contexts
                if "GeneralCeia" in assoc_exclusions and \
                   (time_slot_context.startswith("AntesCeia") or time_slot_context.startswith("DepoisCeia")):
                     if position_name in assoc_exclusions["GeneralCeia"]:
                         is_eligible = False
                     elif position_conceptual_group and any(p in assoc_exclusions["GeneralCeia"] for p in CONCEPTUAL_POSITION_GROUPS.get(position_conceptual_group, [])):
                         is_eligible = False
                # Apply GeneralDraw rules for GeneralDraw context
                elif "GeneralDraw" in assoc_exclusions and time_slot_context == "GeneralDraw":
                     if position_name in assoc_exclusions["GeneralDraw"]:
                         is_eligible = False
                     elif position_conceptual_group and any(p in assoc_exclusions["GeneralDraw"] for p in CONCEPTUAL_POSITION_GROUPS.get(position_conceptual_group, [])):
                         is_eligible = False

        # Apply additional dynamic exclusions (e.g., based on prior allocations in the *same* shift phase)
        if is_eligible and additional_exclusions_for_assoc and assoc in additional_exclusions_for_assoc:
            if position_name in additional_exclusions_for_assoc[assoc]:
                is_eligible = False
            # Check for conceptual group exclusion in dynamic exclusions too
            elif position_conceptual_group and any(p in additional_exclusions_for_assoc[assoc] for p in CONCEPTUAL_POSITION_GROUPS.get(position_conceptual_group, [])):
                is_eligible = False

        # If after ALL exclusions, the associate is still eligible, then apply weights
        if is_eligible:
            # Apply probability weights
            if assoc in probabilities:
                assoc_probabilities = probabilities[assoc]
                # Check specific time slot probability
                if time_slot_context in assoc_probabilities:
                    if position_name in assoc_probabilities[time_slot_context]:
                        current_weight *= assoc_probabilities[time_slot_context][position_name]
                    # Check conceptual group probability for the specific time slot
                    elif position_conceptual_group:
                        # Find the highest weight for any position in the group
                        group_weights = [
                            assoc_probabilities[time_slot_context][p]
                            for p in CONCEPTUAL_POSITION_GROUPS.get(position_conceptual_group, [])
                            if p in assoc_probabilities[time_slot_context]
                        ]
                        if group_weights:
                            current_weight *= max(group_weights) # Apply the strongest probability from the group
                
                # Check general context probability (GeneralCeia/GeneralDraw)
                elif "GeneralCeia" in assoc_probabilities and \
                     (time_slot_context.startswith("AntesCeia") or time_slot_context.startswith("DepoisCeia")):
                    if position_name in assoc_probabilities["GeneralCeia"]:
                        current_weight *= assoc_probabilities["GeneralCeia"][position_name]
                    elif position_conceptual_group:
                        group_weights = [
                            assoc_probabilities["GeneralCeia"][p]
                            for p in CONCEPTUAL_POSITION_GROUPS.get(position_conceptual_group, [])
                            if p in assoc_probabilities["GeneralCeia"]
                        ]
                        if group_weights:
                            current_weight *= max(group_weights)
                
                elif "GeneralDraw" in assoc_probabilities and time_slot_context == "GeneralDraw":
                    if position_name in assoc_probabilities["GeneralDraw"]:
                        current_weight *= assoc_probabilities["GeneralDraw"][position_name]
                    elif position_conceptual_group:
                        group_weights = [
                            assoc_probabilities["GeneralDraw"][p]
                            for p in CONCEPTUAL_POSITION_GROUPS.get(position_conceptual_group, [])
                            if p in assoc_probabilities["GeneralDraw"]
                        ]
                        if group_weights:
                            current_weight *= max(group_weights)
            
            eligible_associates.append(assoc)
            weights.append(current_weight)

    if not eligible_associates:
        return None # No eligible associate for this position

    chosen = random.choices(eligible_associates, weights=weights, k=1)[0]
    return chosen

# --- Allocation Function (Before/After Dinner Shifts) ---
def allocate_dinner_shifts(associates, exclusions, increased_probability, core_associates_for_dinner):
    
    allocated_schedule = {}
    associate_ceia_slots_count = collections.defaultdict(int) 
    antes_ceia_allocations_info = {} 

    shuffled_core_positions = list(CORE_POSITIONS)
    random.shuffle(shuffled_core_positions)

    # --- Phase 1: Allocate AntesCeia positions ---
    antes_ceia_positions_to_fill = [f"AntesCeia - {p}" for p in shuffled_core_positions]
    available_for_antesceia = list(associates) 
    
    # Prioritize core associates for AntesCeia slots
    for core_assoc in core_associates_for_dinner:
        if core_assoc in available_for_antesceia and associate_ceia_slots_count[core_assoc] < 1:
            chosen_position_for_core = None
            
            temp_core_pos_for_antes = list(antes_ceia_positions_to_fill) 
            random.shuffle(temp_core_pos_for_antes)

            for pos_full_name in temp_core_pos_for_antes:
                time_slot, position_name = pos_full_name.split(" - ")
                if not allocated_schedule.get(pos_full_name): 
                    # CRITICAL CALL: Check if the core associate is eligible for this specific position
                    eligible_check = choose_associate_with_rules(
                        [core_assoc], position_name, time_slot, exclusions, increased_probability
                    )
                    if eligible_check == core_assoc:
                        chosen_position_for_core = pos_full_name
                        break
            
            if chosen_position_for_core:
                allocated_schedule[chosen_position_for_core] = core_assoc
                associate_ceia_slots_count[core_assoc] += 1
                
                antes_ceia_allocations_info[core_assoc] = {
                    "exact_position": position_name,
                    "conceptual_group": _INVERTED_CONCEPTUAL_GROUPS.get(position_name, position_name)
                }
                
                available_for_antesceia.remove(core_assoc) 
                antes_ceia_positions_to_fill.remove(chosen_position_for_core) 
            else:
                pass 


    # Fill remaining AntesCeia positions with other available associates
    for full_pos_str in antes_ceia_positions_to_fill:
        if not allocated_schedule.get(full_pos_str): 
            time_slot, position_name = full_pos_str.split(" - ")
            chosen_associate = choose_associate_with_rules(
                available_for_antesceia, position_name, time_slot, exclusions, increased_probability
            )
            if chosen_associate:
                allocated_schedule[full_pos_str] = chosen_associate
                associate_ceia_slots_count[chosen_associate] += 1
                
                antes_ceia_allocations_info[chosen_associate] = {
                    "exact_position": position_name,
                    "conceptual_group": _INVERTED_CONCEPTUAL_GROUPS.get(position_name, position_name)
                }
                
                available_for_antesceia.remove(chosen_associate) 
            else:
                allocated_schedule[full_pos_str] = "(Vazio/Nenhum Associado Elegível)"


    # --- Phase 2: Allocate DepoisCeia positions ---
    depois_ceia_positions_to_fill = [f"DepoisCeia - {p}" for p in shuffled_core_positions]
    
    associates_eligible_for_depois_ceia = []
    for assoc in associates:
        target_slots = 2 if assoc in core_associates_for_dinner else 1
        if associate_ceia_slots_count[assoc] < target_slots:
            associates_eligible_for_depois_ceia.append(assoc)
    
    random.shuffle(associates_eligible_for_depois_ceia) 

    for full_pos_str in depois_ceia_positions_to_fill:
        if not allocated_schedule.get(full_pos_str):
            time_slot, position_name = full_pos_str.split(" - ")
            
            dynamic_exclusions_for_this_slot = collections.defaultdict(list)
            
            for assoc_to_consider in associates_eligible_for_depois_ceia:
                if assoc_to_consider in antes_ceia_allocations_info:
                    antes_ceia_info = antes_ceia_allocations_info[assoc_to_consider]
                    
                    dynamic_exclusions_for_this_slot[assoc_to_consider].append(antes_ceia_info["exact_position"])
                    
                    conceptual_group_name = antes_ceia_info["conceptual_group"]
                    if conceptual_group_name in CONCEPTUAL_POSITION_GROUPS:
                        for pos_in_group in CONCEPTUAL_POSITION_GROUPS[conceptual_group_name]:
                            if pos_in_group not in dynamic_exclusions_for_this_slot[assoc_to_consider]:
                                dynamic_exclusions_for_this_slot[assoc_to_consider].append(pos_in_group)

            chosen_associate = choose_associate_with_rules(
                associates_eligible_for_depois_ceia, position_name, time_slot, exclusions, increased_probability,
                additional_exclusions_for_assoc=dynamic_exclusions_for_this_slot
            )
            
            if chosen_associate:
                allocated_schedule[full_pos_str] = chosen_associate
                associate_ceia_slots_count[chosen_associate] += 1
                associates_eligible_for_depois_ceia.remove(chosen_associate)
            else:
                allocated_schedule[full_pos_str] = "(Vazio/Nenhum Associado Elegível)"
    
    unallocated_associates_for_next_phase = []
    for assoc in associates:
        target_slots = 2 if assoc in core_associates_for_dinner else 1
        if associate_ceia_slots_count[assoc] < target_slots:
            unallocated_associates_for_next_phase.append(assoc)

    return allocated_schedule, unallocated_associates_for_next_phase

# --- Drawing Function (Including Activated Functions) ---
def draw_activated_functions(associates, exclusions, increased_probability, activate_extras_flag, num_draws_input):
    activate_extras = activate_extras_flag
    num_draws = num_draws_input

    if not activate_extras or num_draws <= 0:
        return {}

    drawn_assignments = {} 
    available_associates_for_draw = list(associates) 
    conceptual_role_pool_for_draw = list(ACTIVATED_FUNCTIONS) # Use this as the pool of functions to draw from
    random.shuffle(conceptual_role_pool_for_draw) 

    for i in range(num_draws):
        chosen_associate = None
        chosen_conceptual_role = None # Now this will be an actual function name

        if not conceptual_role_pool_for_draw: # If no more functions to draw
            drawn_assignments[f"Sorteio Posição {i+1} (Nenhuma Função Disponível)"] = "(Vazio)"
            continue # Move to next draw, but it will also be empty if no associates

        chosen_conceptual_role = conceptual_role_pool_for_draw.pop(0) # Take one function from the shuffled list
        
        chosen_associate = choose_associate_with_rules(
            available_associates_for_draw, chosen_conceptual_role, "GeneralDraw", exclusions, increased_probability
        )

        if chosen_associate:
            drawn_assignments[f"{chosen_conceptual_role}"] = chosen_associate # Store only the function name as key for simplicity
            available_associates_for_draw.remove(chosen_associate)
        else:
            drawn_assignments[f"{chosen_conceptual_role}"] = "(Vazio/Nenhum Associado Elegível)"

        # If no more associates available, fill remaining draw slots as empty
        if not available_associates_for_draw and i < num_draws - 1:
            for j in range(i + 1, num_draws):
                if conceptual_role_pool_for_draw:
                    remaining_role = conceptual_role_pool_for_draw.pop(0)
                    drawn_assignments[f"{remaining_role}"] = "(Vazio)"
                else:
                    drawn_assignments[f"Sorteio Posição {j+1} (Nenhuma Função Disponível)"] = "(Vazio)"
            break # Exit the loop as no more associates to assign
            
    return drawn_assignments

# --- Função para gerar o Excel em memória ---
def generate_excel_in_memory(allocated_schedule, drawn_assignments, all_unallocated_associates, model_workbook_path="modelo_escala.xlsx"):
    try:
        workbook = load_workbook(model_workbook_path)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Escala de Alocação"
        sheet['A1'] = "Posição"
        sheet['B1'] = "Antes Ceia"
        sheet['C1'] = "Depois Ceia"
    
    sheet = workbook.active
    
    # Mapeamento de células do Excel para as posições de ceia
    ceia_mapping = {
        'Recirculação 1': {'AntesCeia': 'B2', 'DepoisCeia': 'C2'},
        'Recirculação 2': {'AntesCeia': 'B3', 'DepoisCeia': 'C3'},
        'Recirculação 3': {'AntesCeia': 'B4', 'DepoisCeia': 'C4'},
        'Pesado 1': {'AntesCeia': 'B5', 'DepoisCeia': 'C5'},
        'Pesado 2': {'AntesCeia': 'B6', 'DepoisCeia': 'C6'},
        'Auditoria 1': {'AntesCeia': 'B7', 'DepoisCeia': 'C7'},
        'Auditoria 2': {'AntesCeia': 'B8', 'DepoisCeia': 'C8'},
        'Auditoria 3': {'AntesCeia': 'B9', 'DepoisCeia': 'C9'},
        'Azul 1': {'AntesCeia': 'B10', 'DepoisCeia': 'C10'},
        'Azul 2': {'AntesCeia': 'B11', 'DepoisCeia': 'C11'},
        'Shuttle 1': {'AntesCeia': 'B13', 'DepoisCeia': 'C13'},
        'Shuttle 2': {'AntesCeia': 'B14', 'DepoisCeia': 'C14'},
        'Shuttle 3': {'AntesCeia': 'B15', 'DepoisCeia': 'C15'},
        'Carregamento 1': {'AntesCeia': 'B16', 'DepoisCeia': 'C16'},
        'Carregamento 2': {'AntesCeia': 'B17', 'DepoisCeia': 'C17'},
        'GAP 1': {'AntesCeia': 'B18', 'DepoisCeia': 'C18'},
        'GAP 2': {'AntesCeia': 'B19', 'DepoisCeia': 'C19'}
    }

    # Limpar células de resultados anteriores (se houver)
    for row_num in range(2, 20): # Limpa de B2:C19 (áreas da ceia)
        sheet[f'B{row_num}'] = ""
        sheet[f'C{row_num}'] = ""
    
    # Limpar funções ativas sorteadas - ATUALIZADO para Coluna G, linha 2
    funcoes_ativas_col = 'G' 
    funcoes_ativas_start_row = 2 
    for row_num in range(funcoes_ativas_start_row, funcoes_ativas_start_row + len(ACTIVATED_FUNCTIONS) + 5): # Limpa um pouco além do esperado
        sheet[f'{funcoes_ativas_col}{row_num}'] = ""
    
    # Limpar não alocados - ATUALIZADO para Coluna F, linha 2
    unallocated_text_col = 'F'
    unallocated_text_start_row = 2 
    for row_num in range(unallocated_text_start_row, unallocated_text_start_row + 30): 
        sheet[f'{unallocated_text_col}{row_num}'] = ""
    
    # Escrever alocações de ceia
    for full_pos_str, associate in allocated_schedule.items():
        time_slot, position_name = full_pos_str.split(" - ")
        if position_name in ceia_mapping and time_slot in ceia_mapping[position_name]:
            cell = ceia_mapping[position_name][time_slot]
            sheet[cell] = associate

    # Escrever atribuições do sorteio - ATUALIZADO
    current_active_row = funcoes_ativas_start_row # Começa na linha 2 na coluna G
    
    # Sort the drawn_assignments by the position name from ACTIVATED_FUNCTIONS list order
    # This ensures a consistent order in the Excel file
    sorted_drawn_assignments_for_excel = []
    for func_name in ACTIVATED_FUNCTIONS:
        if func_name in drawn_assignments:
            sorted_drawn_assignments_for_excel.append((func_name, drawn_assignments[func_name]))
        # Handle cases where a function was requested but not assigned (e.g., "(Vazio)")
        elif f"Sorteio Posição {ACTIVATED_FUNCTIONS.index(func_name) + 1} ({func_name})" in drawn_assignments:
             sorted_drawn_assignments_for_excel.append((func_name, drawn_assignments[f"Sorteio Posição {ACTIVATED_FUNCTIONS.index(func_name) + 1} ({func_name})"]))
        # Also need to capture 'Nenhuma Função Disponível' if it was added
        # This part might need further refinement based on exact keys in drawn_assignments if they are not just function names
        # For now, let's assume drawn_assignments has function names as keys.

    # Re-iterate on drawn_assignments to ensure all are caught, especially if keys are like "Sorteio Posição X (...)"
    # We want "Função: Associado" format.
    final_drawn_output = {}
    for key, value in drawn_assignments.items():
        if key in ACTIVATED_FUNCTIONS: # If key is directly the function name
            final_drawn_output[key] = value
        elif "Sorteio Posição" in key and "(" in key and ")" in key: # If key is "Sorteio Posição X (Function Name)"
            func_name_in_key = key.split('(')[1][:-1]
            final_drawn_output[func_name_in_key] = value
        else: # Fallback for unexpected keys
            final_drawn_output[key] = value

    # Now sort this final_drawn_output based on ACTIVATED_FUNCTIONS order
    sorted_drawn_assignments_for_excel = []
    for func_name in ACTIVATED_FUNCTIONS:
        if func_name in final_drawn_output:
            sorted_drawn_assignments_for_excel.append((func_name, final_drawn_output[func_name]))

    # Add any remaining keys that might not be in ACTIVATED_FUNCTIONS (e.g., "Sorteio Posição X (Nenhuma Função Disponível)")
    for key, value in final_drawn_output.items():
        if not any(key == func[0] for func in sorted_drawn_assignments_for_excel): # If not already added
            sorted_drawn_assignments_for_excel.append((key, value))

    for func_name, associate in sorted_drawn_assignments_for_excel:
        if associate.startswith('('): # Ex: "(Vazio/Nenhum Associado Elegível)"
            sheet[f'{funcoes_ativas_col}{current_active_row}'] = f"{func_name}: {associate}"
        else:
            sheet[f'{funcoes_ativas_col}{current_active_row}'] = f"{func_name}: {associate}"
        current_active_row += 1

    # Escrever associados não alocados/sorteados - ATUALIZADO
    current_unallocated_row = unallocated_text_start_row # Começa na linha 2 na coluna F
    for assoc in all_unallocated_associates:
        sheet[f'{unallocated_text_col}{current_unallocated_row}'] = f"- {assoc}"
        current_unallocated_row += 1

    # Salva o workbook em um buffer de bytes
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0) # Retorna o ponteiro para o início do buffer
    return buffer.getvalue() # Retorna os bytes do arquivo Excel

# --- Streamlit App Interface (Mantido como estava) ---
st.set_page_config(page_title="Alocador de Escalas", page_icon="📊", layout="centered")

st.title("📊 Alocador Automático de Escalas e Sorteios")
st.markdown("Bem-vindo ao seu assistente de alocação de equipes!")
st.markdown("---")

# --- 1. Carregar Associados ---
st.header("1. Carregar Lista de Associados")
st.info("Por favor, carregue um arquivo `.txt` contendo os nomes dos associados, um por linha.")
uploaded_associates_file = st.file_uploader("Arraste ou clique para carregar 'associados.txt'", type="txt", key="associates_uploader")

associates = []
if uploaded_associates_file is not None:
    try:
        raw_associates = [line.strip() for line in uploaded_associates_file.getvalue().decode("utf-8").splitlines() if line.strip()]
        # Remover duplicatas mantendo a ordem original
        seen = set()
        associates = [x for x in raw_associates if not (x in seen or seen.add(x))]
        
        if associates:
            st.success(f"Arquivo 'associados.txt' carregado com sucesso! ({len(associates)} associados)")
            st.write("Associados carregados:", associates)
            st.session_state['initial_associates_set'] = set(associates) # Salva para controle final
        else:
            st.warning("O arquivo 'associados.txt' está vazio ou não contém nomes válidos.")
            st.session_state['initial_associates_set'] = set()
    except Exception as e:
        st.error(f"Erro ao ler 'associados.txt': {e}. Verifique se o formato está correto.")
        st.session_state['initial_associates_set'] = set()
else:
    st.session_state['initial_associates_set'] = set()


st.markdown("---")

# --- 2. Executar Alocação de Ceia ---
st.header("2. Alocação de Escala de Ceia")

if st.button("🚀 Iniciar Alocação de Ceia", disabled=not associates):
    if associates:
        with st.spinner("Alocando posições de ceia..."):
            allocated_schedule, unallocated_from_ceia_for_draw = allocate_dinner_shifts(
                list(associates), EXCLUSIONS, INCREASED_PROBABILITY, CORE_ASSOCIATES_FOR_DINNER
            )
            st.session_state['allocated_schedule'] = allocated_schedule
            st.session_state['unallocated_for_draw'] = unallocated_from_ceia_for_draw

            st.subheader("✅ Escala da Ceia Gerada:")
            ceia_data = []
            
            # Ordena a exibição para o usuário
            def sort_key(item_tuple):
                item_key = item_tuple[0]
                time_slot_part, pos_name_part = item_key.split(" - ")
                time_slot_order = 0 if time_slot_part == "AntesCeia" else 1
                
                pos_order_map = {
                    "Recirculação 1": 0, "Recirculação 2": 1, "Recirculação 3": 2,
                    "Pesado 1": 3, "Pesado 2": 4,
                    "Auditoria 1": 5, "Auditoria 2": 6, "Auditoria 3": 7,
                    "Azul 1": 8, "Azul 2": 9,
                    "Shuttle 1": 10, "Shuttle 2": 11, "Shuttle 3": 12,
                    "Carregamento 1": 13, "Carregamento 2": 14,
                    "GAP 1": 15, "GAP 2": 16
                }
                pos_order = pos_order_map.get(pos_name_part, len(pos_order_map))

                return (time_slot_order, pos_order)

            sorted_schedule_items = sorted(allocated_schedule.items(), key=sort_key)

            for full_pos_str, assoc in sorted_schedule_items:
                ceia_data.append({"Posição na Ceia": full_pos_str.replace("AntesCeia - ", "").replace("DepoisCeia - ", ""), "Associado": assoc})
            
            st.dataframe(pd.DataFrame(ceia_data))
            
            if unallocated_from_ceia_for_draw:
                st.info(f"**{len(unallocated_from_ceia_for_draw)}** associados estão elegíveis para o sorteio de funções extras.")
                st.write("Associados disponíveis para sorteio:", unallocated_from_ceia_for_draw)
            else:
                st.info("Todos os associados elegíveis foram alocados na ceia (ou não há mais elegíveis para sorteio).")
            
            st.success("Alocação de Ceia Concluída! Prossiga para o sorteio, se desejar.")
    else:
        st.warning("Por favor, carregue a lista de associados primeiro.")

st.markdown("---")

# --- 3. Sorteio de Funções Extras ---
st.header("3. Sorteio de Funções Extras")

unallocated_for_draw = st.session_state.get('unallocated_for_draw', [])
max_draws = len(ACTIVATED_FUNCTIONS) # Número máximo de funções ativadas
max_associates_for_draw = len(unallocated_for_draw)

if not unallocated_for_draw:
    st.warning("Nenhum associado disponível para sorteio de funções extras. Execute a alocação de ceia primeiro.")
else:
    activate_extras_option = st.radio(
        "Deseja ativar as funções extras?",
        ("Sim", "Não"), index=1, key="activate_extras_radio"
    )
    
    num_draws = 0
    if activate_extras_option == "Sim":
        st.info(f"Atualmente, há {max_associates_for_draw} associados disponíveis e {max_draws} funções extras possíveis.")
        num_draws = st.number_input(
            f"Quantas funções extras deseja sortear? (Máximo: {min(max_draws, max_associates_for_draw)})",
            min_value=0, max_value=min(max_draws, max_associates_for_draw), value=min(max_draws, max_associates_for_draw), key="num_draws_input"
        )

    if st.button("🎲 Executar Sorteio de Funções Extras", disabled=(activate_extras_option == "Sim" and num_draws == 0) or not unallocated_for_draw, key="draw_button"):
        if activate_extras_option == "Sim" and num_draws > 0:
            with st.spinner("Executando sorteio de funções..."):
                drawn_assignments = draw_activated_functions(
                    list(unallocated_for_draw), # Passa uma cópia para a função
                    EXCLUSIONS, INCREASED_PROBABILITY,
                    True, # activate_extras_flag
                    num_draws # num_draws_input
                )
                st.session_state['drawn_assignments'] = drawn_assignments

                st.subheader("✅ Atribuições do Sorteio Finalizadas:")
                drawn_data = []
                for pos, assoc in sorted(drawn_assignments.items()): # Agora 'pos' será o nome da função
                    drawn_data.append({"Posição Sorteada": pos, "Associado": assoc})
                st.dataframe(pd.DataFrame(drawn_data))
                st.success("Sorteio Concluído!")
        elif activate_extras_option == "Não":
            st.info("Sorteio de funções extras desativado.")
            st.session_state['drawn_assignments'] = {} # Garante que está vazio se não sorteou
        else:
            st.warning("Número de sorteios deve ser maior que zero se ativado.")

st.markdown("---")

# --- 4. Baixar Escala Final ---
st.header("4. Baixar Escala Completa em Excel")

allocated_schedule = st.session_state.get('allocated_schedule', {})
drawn_assignments = st.session_state.get('drawn_assignments', {})
initial_associates_set = st.session_state.get('initial_associates_set', set())

if not allocated_schedule:
    st.warning("Por favor, execute a alocação de ceia primeiro para gerar o Excel.")
else:
    if st.button("💾 Gerar e Baixar Escala Completa em Excel", key="download_excel_button"):
        with st.spinner("Gerando arquivo Excel..."):
            # Obter todos os associados alocados/sorteados para a lista final de não alocados
            final_allocated_associates_set = set()
            for assoc in allocated_schedule.values():
                if not assoc.startswith('('):
                    final_allocated_associates_set.add(assoc)
            for assoc in drawn_assignments.values():
                if not assoc.startswith('('):
                    final_allocated_associates_set.add(assoc)
            
            all_unallocated_associates_overall = list(initial_associates_set - final_allocated_associates_set)
            all_unallocated_associates_overall.sort() # Para manter a ordem

            try:
                excel_bytes = generate_excel_in_memory(
                    allocated_schedule, 
                    drawn_assignments, 
                    all_unallocated_associates_overall,
                    "modelo_escala.xlsx" # Nome do arquivo do modelo que deve estar no GitHub
                )
                
                st.download_button(
                    label="Clique para Baixar 'escala_da_equipe.xlsx'",
                    data=excel_bytes,
                    file_name="escala_da_equipe.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("Arquivo Excel gerado com sucesso!")

                if all_unallocated_associates_overall:
                    st.subheader("Associados Restantes (Não Alocados/Sorteados em Nenhuma Posição):")
                    for assoc in all_unallocated_associates_overall:
                        st.write(f"- {assoc}")
                else:
                    st.info("🎉 Todos os associados foram alocados/sorteados para alguma posição!")

            except Exception as e:
                st.error(f"Ocorreu um erro ao gerar o arquivo Excel: {e}")
                st.warning("Verifique se o 'modelo_escala.xlsx' está no formato correto e na mesma pasta do 'app.py' no seu repositório GitHub.")

st.markdown("---")

# --- Rodapé ---
st.markdown("Esta aplicação foi desenvolvida por **Rinalanc/Github**.")
st.markdown("Data: 30/06/2025")
